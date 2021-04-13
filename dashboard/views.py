import openpyxl
import xlsxwriter
from openpyxl.utils import get_column_letter
from django.contrib.auth import logout, authenticate, login, update_session_auth_hash
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib import messages
from django.contrib.auth.hashers import make_password
from django.core.exceptions import ValidationError
from django.core.files import File
from django.core.files.storage import FileSystemStorage
from django.db import IntegrityError
from django.db.models import Q
from django.http import HttpResponseRedirect, Http404, HttpResponse
from django.shortcuts import render, get_object_or_404, redirect
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from psycopg2 import errorcodes
import random
import io

from django.urls import reverse
from psycopg2.errorcodes import UNIQUE_VIOLATION

from dashboard.forms import *
from dashboard.models import *


def pagination_template(request, source, divide_by=2, page_name='page'):
    page = request.GET.get(page_name, 1)
    paginator = Paginator(source, divide_by)
    try:
        pager = paginator.page(page)
    except PageNotAnInteger:
        pager = paginator.page(1)
    except EmptyPage:
        pager = paginator.page(paginator.num_pages)

    return pager


def pagination_template_long(request, source, divide_by=2, page_name='page'):
    page = request.GET.get(page_name, 1)
    paginator = Paginator(source, divide_by)
    try:
        pager = paginator.page(page)
    except PageNotAnInteger:
        pager = paginator.page(1)
    except EmptyPage:
        pager = paginator.page(paginator.num_pages)

    # Get the index of the current page
    index = pager.number - 1  # edited to something easier without index
    # This value is maximum index of your pages, so the last page - 1
    max_index = len(paginator.page_range)
    # You want a range of 7, so lets calculate where to slice the list
    start_index = index - 3 if index >= 3 else 0
    end_index = index + 3 if index <= max_index - 3 else max_index
    # Get our new page range. In the latest versions of Django page_range returns
    # an iterator. Thus pass it to list, to make our slice possible again.
    page_range = list(paginator.page_range)[start_index:end_index]

    return pager, page_range


def pagination_show(request, all_objects, divide_by=6):
    if len(all_objects) == 0:
        return {
            'color': 'danger',
            'text': ' موردی یافت نشد!',
        }
    else:
        return pagination_template(request, source=all_objects, divide_by=divide_by)


def pagination_show_long(request, all_objects, divide_by=6):
    if len(all_objects) == 0:
        return {
            'color': 'danger',
            'text': ' موردی یافت نشد!',
        }
    else:
        return pagination_template_long(request, source=all_objects, divide_by=divide_by)


def signup_view(request):
    notification = Notification.objects.filter(show_on_signup=True).order_by('-date')
    if request.method == 'POST':
        signup_form = SignUpForm(request.POST)
        if signup_form.is_valid():
            signup_form.save(commit=False)
            if signup_form.cleaned_data.get('user_type') == 1:
                signup_form.instance.is_student = True
            elif signup_form.cleaned_data.get('user_type') == 2:
                signup_form.instance.is_teacher = True
            elif signup_form.cleaned_data.get('user_type') == 3:
                signup_form.instance.is_supervisor = True
            signup_form.save()
            username = signup_form.cleaned_data.get('username')
            raw_password = signup_form.cleaned_data.get('password1')
            user = authenticate(username=username, password=raw_password)
            login(request, user)
            return HttpResponseRedirect(reverse('dashboard:me_view'))
    else:
        signup_form = SignUpForm()
        # signup_form_username = CustomAuthenticationForm
    if request.user.is_authenticated:
        return HttpResponseRedirect(reverse('home_page'))
    context = {
        'signup_form': signup_form,
        'notification': notification,
        # 'signup_form_username': signup_form_username,
    }
    return render(request, 'dashboard/signup.html', context)


def login_view(request):
    notification = Notification.objects.filter(show_on_login=True).order_by('-date')
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user is None:
            context = {
                'username': username,
                'error': 'نام کاربری یا رمز عبور اشتباه است!',
                'notification': notification,
            }
        else:
            login(request, user)
            # return HttpResponseRedirect(reverse('home_page'))
            return HttpResponseRedirect(reverse('dashboard:me_view'))
    elif request.user.is_authenticated:
        # return HttpResponseRedirect(reverse('home_page'))
        return HttpResponseRedirect(reverse('dashboard:me_view'))
    else:
        context = {
            'notification': notification,
        }
    return render(request, 'dashboard/login.html', context)


def logout_view(request):
    logout(request)
    return HttpResponseRedirect(reverse('dashboard:login'))


# def superuser_dashboard(request):
#     temp_user = request.user
#     temp_profile = Profile.objects.get(user=temp_user)
#     # temp_student = Profile.objects.filter(user__is_student=True)
#     # temp_teacher = Profile.objects.filter(user__is_teacher=True)
#     # temp_supervisor = Profile.objects.filter(user__is_supervisor=True)
#     # temp_notification = Notification.objects.all().order_by('-date')
#     # temp_academy = Academy.objects.all()
#     # temp_klass = Klass.objects.all()
#     # temp_exam = Exam.objects.all()
#     # temp_question = Question.objects.all()
#     # temp_score = StudentScore.objects.all()
#     # temp_category = Category.objects.all()
#     return {
#         'is_superuser': True,
#         'user': temp_user,
#         'profile': temp_profile,
#         # 'student': temp_student,
#         # 'teacher': temp_teacher,
#         # 'supervisor': temp_supervisor,
#         # 'notification': temp_notification,
#         # 'academy': temp_academy,
#         # 'klass': temp_klass,
#         # 'exam': temp_exam,
#         # 'question': temp_question,
#         # 'score': temp_score,
#         # 'category': temp_category,
#     }
#
#
# def supervisor_dashboard(request):
#     temp_user = request.user
#     temp_profile = Profile.objects.get(user=temp_user)
#     temp_supervisor = Profile.objects.get(Q(user=temp_user) & Q(user__is_supervisor=True))
#     temp_academy = temp_profile.academy
#     temp_teacher = Profile.objects.filter(Q(user__profile__academy=temp_academy) & Q(user__is_teacher=True))
#     temp_student = Profile.objects.filter(Q(user__profile__academy=temp_academy) & Q(user__is_student=True))
#     temp_klass = Klass.objects.filter(academy=temp_academy)
#     temp_exam = Exam.objects.filter(exam_klass__academy=temp_academy).order_by('-exam_finish')
#     temp_score = StudentScore.objects.filter(student__user__profile__academy=temp_academy).order_by('-exam')
#     temp_wallet = Wallet.objects.get(user=temp_user)
#     temp_transaction = Transaction.objects.filter(wallet__user=temp_user).order_by('-order_time')
#     temp_question = []
#     for i in temp_teacher:
#         x = Question.objects.filter(creator=i)
#         for j in x:
#             temp_question.append(j)
#     return {
#         'is_supervisor': True,
#         'user': temp_user,
#         'profile': temp_profile,
#         'supervisor': temp_supervisor,
#         'teacher': temp_teacher,
#         'student': temp_student,
#         'academy': temp_academy,
#         'klass': temp_klass,
#         'exam': temp_exam,
#         'score': temp_score,
#         'question': temp_question,
#         'wallet': temp_wallet,
#         'transaction': temp_transaction,
#     }
#
#
# def teacher_dashboard(request):
#     temp_user = request.user
#     temp_profile = Profile.objects.get(user=temp_user)
#     temp_teacher = Profile.objects.get(Q(user=temp_user) & Q(user__is_teacher=True))
#     temp_academy = temp_profile.academy
#     temp_klass = Klass.objects.filter(teacher=temp_profile)
#     temp_student = Profile.objects.filter(
#         Q(user__profile__klass__teacher=temp_teacher) & Q(user__is_student=True))
#     temp_exam = Exam.objects.filter(exam_klass__teacher=temp_teacher)
#     temp_question = Question.objects.filter(creator=temp_teacher)
#     temp_score = StudentScore.objects.filter(student__user__profile__klass__teacher=temp_teacher)
#     temp_wallet = Wallet.objects.get(user=temp_user)
#     temp_transaction = Transaction.objects.filter(wallet__user=temp_user).order_by('-order_time')
#     return {
#         'is_teacher': True,
#         'user': temp_user,
#         'profile': temp_profile,
#         'teacher': temp_teacher,
#         'student': temp_student,
#         'academy': temp_academy,
#         'klass': temp_klass,
#         'exam': temp_exam,
#         'question': temp_question,
#         'score': temp_score,
#         'wallet': temp_wallet,
#         'transaction': temp_transaction,
#     }
#
#
# def student_dashboard(request):
#     temp_user = request.user
#     temp_profile = Profile.objects.get(user=temp_user)
#     temp_student = Profile.objects.get(Q(user=temp_user) & Q(user__is_student=True))
#     temp_academy = temp_profile.academy
#     temp_klass = temp_profile.klass.all()
#     try:
#         temp_klass = temp_klass[0]
#     except:
#         temp_klass = None
#     temp_exam = Exam.objects.filter(exam_klass=temp_klass)
#     temp_score = StudentScore.objects.filter(student=temp_student)
#     return {
#         'is_student': True,
#         'user': temp_user,
#         'profile': temp_profile,
#         'student': temp_student,
#         'academy': temp_academy,
#         'klass': temp_klass,
#         'exam': temp_exam,
#         'score': temp_score,
#     }


# def goes_in_all_view(request):
#     temp_profile = Profile.objects.get(user=request.user)
#     temp_dashboard = None
#     if temp_profile.user.is_superuser:
#         temp_dashboard = superuser_dashboard(request)
#     elif temp_profile.user.is_supervisor:
#         temp_dashboard = supervisor_dashboard(request)
#     elif temp_profile.user.is_teacher:
#         temp_dashboard = teacher_dashboard(request)
#     elif temp_profile.user.is_student:
#         temp_dashboard = student_dashboard(request)
#     return {
#         'profile': temp_profile,
#         'dashboard': temp_dashboard,
#     }


@login_required
def me_view(request):
    profile = Profile.objects.get(user=request.user)
    user_type = profile.user_type_finder()
    jalali = profile.show_birth_date_in_jalali()
    notification = show_notification()
    # g_i_a_v = goes_in_all_view(request)
    context = {
        'profile': profile,
        'user_type': user_type,
        'jalali': jalali,
        'notification': notification,
        # 'g_i_a_v': g_i_a_v,
    }
    return render(request, 'dashboard/me_page.html', context)


@login_required
def sign_up_key_view(request):
    profile = request.user.profile
    output_msg, current_key = None, None
    # g_i_a_v = goes_in_all_view(request)
    if request.method == 'POST':
        sign_up_key = SignUpKeyCheckerForm(request.POST)
        if sign_up_key.is_valid():
            if SignUpKey.objects.filter(pk=sign_up_key.cleaned_data['sign_up_key']).count() == 0:
                output_msg = {
                    'color': 'danger',
                    'text': 'کلید ثبت نام صحیح نمی باشد!',
                }
            else:
                temp_kls = SignUpKey.objects.get(pk=sign_up_key.cleaned_data['sign_up_key'])
                temp_kls = temp_kls.klass
                temp_kls2 = profile.klass.all()
                if len(temp_kls2) != 0:
                    for i in temp_kls2:
                        profile.klass.remove(i)
                profile.klass.add(temp_kls)
                profile.academy = temp_kls.academy
                profile.save()
                output_msg = {
                    'color': 'success',
                    'text': 'با موفقیت ثبت شد!',
                }
    else:
        sign_up_key = SignUpKeyCheckerForm()
    if profile.klass:
        temp_kls2 = profile.klass.all()
        if temp_kls2.count() != 0:
            current_key = SignUpKey.objects.get(klass=temp_kls2[0]).pk
    context = {
        'profile': profile,
        # 'g_i_a_v': g_i_a_v,
        'output_msg': output_msg,
        'current_key': current_key,
        'sign_up_key': sign_up_key,
    }
    return render(request, 'dashboard/sign_up_key.html', context)


@login_required
def profile_edit_view(request):
    profile = request.user.profile
    # g_i_a_v = goes_in_all_view(request)
    user_object = request.user
    if request.method == 'POST':
        change_profile = ChangeProfileForm(request.POST, request.FILES, instance=profile)
        change_profile_user_model = ChangeProfileUserModelForm(request.POST, instance=user_object)
        if change_profile.is_valid() and change_profile_user_model.is_valid():
            change_profile.save()
            change_profile_user_model.save()
            return HttpResponseRedirect(reverse('dashboard:me_view'))
    else:
        change_profile = ChangeProfileForm(instance=profile)
        change_profile_user_model = ChangeProfileUserModelForm(instance=user_object)
    context = {
        'profile': profile,
        'change_profile': change_profile,
        'change_profile_user_model': change_profile_user_model,
        # 'g_i_a_v': g_i_a_v,
    }
    return render(request, 'dashboard/profile_edit.html', context)


@login_required
def password_edit_view(request):
    profile = Profile.objects.get(user=request.user)
    # g_i_a_v = goes_in_all_view(request)
    if request.method == 'POST':
        change_password = CustomPasswordChangeForm(request.user, request.POST)
        if change_password.is_valid():
            user = change_password.save()
            update_session_auth_hash(request, user)
            messages.success(request, 'رمز شما با موفقیت بروز شد')
            return HttpResponseRedirect(reverse('dashboard:me_view'))
        else:
            messages.error(request, 'لطفا خطای زیر را بررسی کنید')
    else:
        change_password = CustomPasswordChangeForm(request.user)
    context = {
        'profile': profile,
        'change_password': change_password,
        # 'g_i_a_v': g_i_a_v,
    }
    return render(request, 'dashboard/password_edit.html', context)


def is_user_registered(request):
    profile = None
    if request.user.is_authenticated:
        profile = Profile.objects.get(user=request.user)
    return profile


def home_page_view(request):
    temp_var = is_user_registered(request)
    notification = show_notification()
    # g_i_a_v = None
    # if temp_var:
    # g_i_a_v = goes_in_all_view(request)
    context = {
        'profile': temp_var,
        'notification': notification,
        # 'g_i_a_v': g_i_a_v,
        'home_page_statics': {
            'all_students': Profile.objects.filter(user__is_student=True).count(),
            'all_exams': Exam.objects.all().count(),
            'all_academies': Academy.objects.all().count(),
            'all_questions': Question.objects.all().count(),
            'all_teachers': Profile.objects.filter(user__is_teacher=True).count(),
            'all_users': User.objects.all().count(),
        }
    }
    return render(request, 'dashboard/home_page.html', context)


# //////////////////////////////////////////// CATEGORIES ////////////////////////////////////////////
@login_required
@permission_required('dashboard.view_category', raise_exception=True)
def category_create_view(request):
    # g_i_a_v = goes_in_all_view(request)

    if request.method == 'POST':
        form = ChangeCategoryForm(request.POST)
        if form.is_valid():
            form.save()
            output_msg = {
                'color': 'success',
                'text': 'دسته بندی با موفقیت ایجاد شد!'
            }
        else:
            output_msg = {
                'color': 'danger',
                'text': 'متاسفانه ایجاد دسته بندی با مشکل مواجه شد!'
            }
    else:
        form = ChangeCategoryForm()
        output_msg = None
    context = {
        'page_name': 'دسته بندی',
        # 'g_i_a_v': g_i_a_v,
        'form': form,
        'output_msg': output_msg,
        'profile': Profile.objects.get(user=request.user),
    }
    return render(request, 'dashboard/category/category_create.html', context)


@login_required
@permission_required(['dashboard.view_category', 'dashboard.change_category', 'dashboard.delete_category'],
                     raise_exception=True)
def category_manage_view(request):
    # g_i_a_v = goes_in_all_view(request)
    all_object = Category.objects.all()
    if all_object.count() == 0:
        output_msg = {
            'color': 'danger',
            'text': 'دسته بندی ای برای نمایش وجود ندارد!'
        }
    else:
        output_msg = None
    all_object = pagination_show(request, all_objects=all_object)
    context = {
        'page_name': 'دسته بندی',
        # 'g_i_a_v': g_i_a_v,
        'all_object': all_object,
        'output_msg': output_msg,
        'profile': Profile.objects.get(user=request.user),
    }
    return render(request, 'dashboard/superuser_dashboard/category_manage.html', context)


@login_required
@permission_required(['dashboard.view_category', 'dashboard.change_category', 'dashboard.delete_category'],
                     raise_exception=True)
def category_change_view(request, cat_id):
    # g_i_a_v = goes_in_all_view(request)
    category = Category.objects.get(pk=cat_id)
    if request.method == 'POST':
        form = ChangeCategoryForm(request.POST, instance=category)
        if form.is_valid():
            form.save()
            output_msg = {
                'color': 'success',
                'text': 'دسته بندی با موفقیت ویرایش شد!'
            }
        else:
            output_msg = {
                'color': 'danger',
                'text': 'متاسفانه ویرایش دسته بندی با مشکل مواجه شد!'
            }
    else:
        form = ChangeCategoryForm(instance=category)

        output_msg = None
    context = {
        'page_name': 'دسته بندی',
        # 'g_i_a_v': g_i_a_v,
        'output_msg': output_msg,
        'form': form,
        'profile': Profile.objects.get(user=request.user),
    }
    return render(request, 'dashboard/notification/notification_change.html', context)


# //////////////////////////////////////////// NOTIFICATIONS ////////////////////////////////////////////
@login_required
@permission_required('dashboard.view_notification', raise_exception=True)
def notification_create_view(request):
    # g_i_a_v = goes_in_all_view(request)

    if request.method == 'POST':
        form = NotificationChangeForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            output_msg = {
                'color': 'success',
                'text': 'اطلاعیه با موفقیت ایجاد شد!'
            }
        else:
            output_msg = {
                'color': 'danger',
                'text': 'متاسفانه ایجاد اطلاعیه با مشکل مواجه شد!'
            }
    else:
        form = NotificationChangeForm()
        output_msg = None
    context = {
        'page_name': 'اطلاعیه',
        # 'g_i_a_v': g_i_a_v,
        'form': form,
        'output_msg': output_msg,
        'profile': Profile.objects.get(user=request.user),
    }
    return render(request, 'dashboard/notification/notification_create.html', context)


@login_required
@permission_required(['dashboard.view_notification', 'dashboard.change_notification', 'dashboard.delete_notification'],
                     raise_exception=True)
def notification_manage_view(request):
    # g_i_a_v = goes_in_all_view(request)
    all_object = Notification.objects.all()
    if all_object.count() == 0:
        output_msg = {
            'color': 'danger',
            'text': 'اطلاعیه ای برای نمایش وجود ندارد!'
        }
    else:
        output_msg = None
    all_object = pagination_show(request, all_objects=all_object)
    context = {
        'page_name': 'اطلاعیه',
        # 'g_i_a_v': g_i_a_v,
        'all_object': all_object,
        'output_msg': output_msg,
        'profile': Profile.objects.get(user=request.user),
    }
    return render(request, 'dashboard/superuser_dashboard/notification_manage.html', context)


@login_required
@permission_required(['dashboard.view_notification', 'dashboard.change_notification', 'dashboard.delete_notification'],
                     raise_exception=True)
def notification_change_view(request, ntf_id):
    # g_i_a_v = goes_in_all_view(request)
    notification = Notification.objects.get(pk=ntf_id)
    if request.method == 'POST':
        form = NotificationChangeForm(request.POST, request.FILES, instance=notification)
        if form.is_valid():
            form.save()
            output_msg = {
                'color': 'success',
                'text': 'اطلاعیه با موفقیت ویرایش شد!'
            }
        else:
            output_msg = {
                'color': 'danger',
                'text': 'متاسفانه ویرایش اطلاعیه با مشکل مواجه شد!'
            }
    else:
        form = NotificationChangeForm(instance=notification)

        output_msg = None
    context = {
        'page_name': 'اطلاعیه',
        # 'g_i_a_v': g_i_a_v,
        'output_msg': output_msg,
        'form': form,
        'profile': Profile.objects.get(user=request.user),
    }
    return render(request, 'dashboard/notification/notification_change.html', context)


# //////////////////////////////////////////// ACADEMIES ////////////////////////////////////////////
@login_required
@permission_required('dashboard.add_academy', raise_exception=True)
def academy_create_view(request):
    # g_i_a_v = goes_in_all_view(request)

    if request.method == 'POST':
        form = AcademyChangeForm(request.POST)
        all_supervisors = SupervisorFinderForm(request.POST)
        if form.is_valid() and all_supervisors.is_valid():
            temp_aca = form.save(commit=False)
            temp_profile = all_supervisors.cleaned_data['supervisor']
            temp_profile.academy = temp_aca
            form.save()
            temp_profile.save()
            output_msg = {
                'color': 'success',
                'text': 'آموزشگاه با موفقیت ایجاد شد!'
            }
        else:
            output_msg = {
                'color': 'danger',
                'text': 'متاسفانه ایجاد آموزشگاه با مشکل مواجه شد!'
            }
    else:
        form = AcademyChangeForm()
        all_supervisors = SupervisorFinderForm()
        output_msg = None
    context = {
        'page_name': 'آموزشگاه',
        # 'g_i_a_v': g_i_a_v,
        'all_supervisors': all_supervisors,
        'form': form,
        'output_msg': output_msg,
        'profile': Profile.objects.get(user=request.user),
    }
    return render(request, 'dashboard/academy/academy_create.html', context)


@login_required
@permission_required(['dashboard.view_academy', 'dashboard.change_academy', 'dashboard.delete_academy'],
                     raise_exception=True)
def academy_manage_view(request):
    # g_i_a_v = goes_in_all_view(request)
    profile = Profile.objects.get(user=request.user)
    all_object = Academy.objects.all()
    if request.method == 'POST':
        if request.POST.get('search_content'):
            all_object = Academy.objects.all()
            all_object = all_object.filter(name__contains=request.POST.get('search_content'))
    if all_object.count() == 0:
        output_msg = {
            'color': 'danger',
            'text': 'آموزشگاهی برای نمایش وجود ندارد!'
        }
    else:
        output_msg = None
    all_object, page_range = pagination_template_long(request, all_object, divide_by=6)
    context = {
        'page_name': 'آموزشگاه',
        # 'g_i_a_v': g_i_a_v,
        'all_object': all_object,
        'output_msg': output_msg,
        'page_range': page_range,
        'profile': profile,
    }
    return render(request, 'dashboard/superuser_dashboard/academy_manage.html', context)


@login_required
@permission_required(['dashboard.view_academy', 'dashboard.change_academy', 'dashboard.delete_academy'],
                     raise_exception=True)
def academy_change_view(request, aca_id):
    # g_i_a_v = goes_in_all_view(request)
    academy = Academy.objects.get(pk=aca_id)
    manager = Profile.objects.filter(Q(academy=academy) & Q(user__is_supervisor=True))
    if manager.count() == 0:
        manager = 'موردی ثبت نشده!'
    else:
        manager = manager[0].user.get_full_name()
    if request.method == 'POST':
        form = AcademyChangeForm(request.POST, instance=academy)
        all_supervisors = SupervisorFinderForm(request.POST)
        if form.is_valid() and all_supervisors.is_valid():
            temp_aca = form.save(commit=False)
            temp_profile = all_supervisors.cleaned_data['supervisor']
            temp_profile.academy = temp_aca
            form.save()
            temp_profile.save()
            output_msg = {
                'color': 'success',
                'text': 'آموزشگاه با موفقیت ویرایش شد!'
            }
        else:
            output_msg = {
                'color': 'danger',
                'text': 'متاسفانه ویرایش آموزشگاه با مشکل مواجه شد!'
            }
    else:
        form = AcademyChangeForm(instance=academy)
        all_supervisors = SupervisorFinderForm()
        output_msg = None
    context = {
        'page_name': 'آموزشگاه',
        # 'g_i_a_v': g_i_a_v,
        'output_msg': output_msg,
        'manager': manager,
        'all_supervisors': all_supervisors,
        'form': form,
        'profile': Profile.objects.get(user=request.user),
    }
    return render(request, 'dashboard/academy/academy_change.html', context)


@login_required
def academy_view_view(request, aca_id):
    # g_i_a_v = goes_in_all_view(request)
    academy = get_object_or_404(Academy, pk=aca_id)
    academy_people = Profile.objects.filter(academy=academy)
    academy_teacher = Profile.objects.filter(Q(academy=academy) & Q(user__is_teacher=True))
    academy_klass = Klass.objects.filter(academy=academy)
    manager = Profile.objects.filter(Q(academy=academy) & Q(user__is_supervisor=True))
    if manager.count() == 0:
        manager = 'موردی ثبت نشده!'
    else:
        manager = manager[0].user.get_full_name()
    context = {
        # 'g_i_a_v': g_i_a_v,
        'profile': Profile.objects.get(user=request.user),
        'academy': academy,
        'manager': manager,
        'academy_klass': academy_klass,
        'academy_people': academy_people,
        'academy_teacher': academy_teacher,
    }
    return render(request, 'dashboard/academy/academy_dashboard.html', context)


# //////////////////////////////////////////// KLASSES ////////////////////////////////////////////
@login_required
@permission_required('dashboard.add_klass', raise_exception=True)
def klass_create_view(request):
    # g_i_a_v = goes_in_all_view(request)

    if request.method == 'POST':
        form = KlassChangeForm(request.POST)
        if form.is_valid():
            temp_kls = form.save()
            temp_teacher = temp_kls.teacher.all()
            if len(temp_teacher) != 0:
                for i in temp_teacher:
                    i.academy = temp_kls.academy
                    i.klass.add(temp_kls)
                    i.save()
            output_msg = {
                'color': 'success',
                'text': 'کلاس با موفقیت ایجاد شد!'
            }
        else:
            output_msg = {
                'color': 'danger',
                'text': 'متاسفانه ایجاد کلاس با مشکل مواجه شد!'
            }
    else:
        form = KlassChangeForm()
        output_msg = None
    context = {
        'page_name': 'کلاس',
        # 'g_i_a_v': g_i_a_v,
        'form': form,
        'output_msg': output_msg,
        'profile': Profile.objects.get(user=request.user),
    }
    return render(request, 'dashboard/klass/klass_create.html', context)


@login_required
@permission_required(['dashboard.add_question'], raise_exception=True)
def klass_manage_view(request):
    # g_i_a_v = goes_in_all_view(request)
    profile = Profile.objects.get(user=request.user)
    if profile.user.is_superuser:
        all_object = Klass.objects.all()
    else:
        all_object = Klass.objects.filter(academy=profile.academy)
    if request.method == 'POST':
        if request.POST.get('search_content'):
            all_object = all_object
            all_object = all_object.filter(Q(name__contains=request.POST.get('search_content')) | Q(
                academy__name__contains=request.POST.get('search_content')))
    if all_object.count() == 0:
        output_msg = {
            'color': 'danger',
            'text': 'کلاسی برای نمایش وجود ندارد!'
        }
    else:
        output_msg = None
    all_object, page_range = pagination_template_long(request, all_object, divide_by=6)
    context = {
        'page_name': 'کلاس',
        # 'g_i_a_v': g_i_a_v,
        'all_object': all_object,
        'output_msg': output_msg,
        'page_range': page_range,
        'profile': profile,
    }
    return render(request, 'dashboard/superuser_dashboard/klass_manage.html', context)


@login_required
@permission_required(['dashboard.view_klass', 'dashboard.change_klass', 'dashboard.delete_klass'], raise_exception=True)
def klass_change_view(request, kls_id):
    # g_i_a_v = goes_in_all_view(request)
    klass = Klass.objects.get(pk=kls_id)
    if request.method == 'POST':
        form = KlassChangeForm(request.POST, instance=klass)
        if form.is_valid():
            temp_kls = form.save()
            temp_teacher = temp_kls.teacher.all()
            if len(temp_teacher) != 0:
                for i in temp_teacher:
                    i.academy = temp_kls.academy
                    i.klass.add(temp_kls)
                    i.save()
            output_msg = {
                'color': 'success',
                'text': 'کلاس با موفقیت ویرایش شد!'
            }
        else:
            output_msg = {
                'color': 'danger',
                'text': 'متاسفانه ویرایش کلاس با مشکل مواجه شد!'
            }
    else:
        form = KlassChangeForm(instance=klass)
        output_msg = None
    context = {
        'page_name': 'کلاس',
        # 'g_i_a_v': g_i_a_v,
        'output_msg': output_msg,
        'form': form,
        'profile': Profile.objects.get(user=request.user),
    }
    return render(request, 'dashboard/klass/klass_change.html', context)


@login_required
def klass_view_view(request, kls_id):
    # g_i_a_v = goes_in_all_view(request)
    klass = get_object_or_404(Klass, pk=kls_id)
    try:
        sign_up_key = SignUpKey.objects.get(klass=klass)
    except:
        while True:
            temp_key = get_random_string(8)
            if SignUpKey.objects.filter(sign_up_key__exact=temp_key).count() == 0:
                sign_up_key = SignUpKey.objects.create(sign_up_key=temp_key, klass=klass)
                break
    klass_student = Profile.objects.filter(Q(user__profile__klass=klass) & Q(user__is_student=True))
    klass_teacher = Profile.objects.filter(Q(user__profile__klass=klass) & Q(user__is_teacher=True))
    context = {
        # 'g_i_a_v': g_i_a_v,
        'klass': klass,
        'klass_student': klass_student,
        'klass_teacher': klass_teacher,
        'sign_up_key': sign_up_key,
        'profile': Profile.objects.get(user=request.user),
    }
    return render(request, 'dashboard/klass/klass_dashboard.html', context)


# //////////////////////////////////////////// QUESTIONS ////////////////////////////////////////////
@login_required
@permission_required('dashboard.add_question', raise_exception=True)
def question_create_view(request):
    # g_i_a_v = goes_in_all_view(request)
    profile = Profile.objects.get(user=request.user)
    output_msg, validation_error = None, None
    if request.method == 'POST':
        question = QuestionCreateForm(request.POST, request.FILES)
        ch = ChoiceFormset(request.POST)
        if question.is_valid():
            question.save(commit=False)
            question.instance.creator = profile
            t_q = question.save()  # t_q:temporary question
            how_many_true = 0
            if ch.is_valid():
                temp = 1
                try:
                    for i in ch:
                        i.save(commit=False)
                        i.instance.choice_position = temp
                        temp += 1
                        i.instance.choice_question = t_q
                        if i.instance.is_correct:
                            how_many_true += 1
                        i.save()
                except IntegrityError as e:
                    if e.__cause__.pgcode == errorcodes.UNIQUE_VIOLATION:
                        validation_error = {
                            'color': 'danger',
                            'text': 'گزینه ها نباید مانند یکدیگر باشند!'
                        }
            else:
                how_many_true = 1
            if how_many_true != 1:
                for i in ch:
                    i.instance.delete()
                t_q.delete()
                output_msg = {
                    'color': 'danger',
                    'text': 'تعداد گزینه های صحیح ایراد دارد!'
                }
            else:
                output_msg = {
                    'color': 'success',
                    'text': 'سوال با موفقیت ایجاد شد!'
                }
        else:
            output_msg = {
                'color': 'danger',
                'text': 'متاسفانه ایجاد سوال با مشکل مواجه شد!'
            }
    else:
        question = QuestionCreateForm()
        ch = ChoiceFormset()
    if validation_error:
        output_msg = validation_error
    context = {
        'page_name': 'سوال',
        # 'g_i_a_v': g_i_a_v,
        'profile': profile,
        'question': question,
        'ch': ch,
        'output_msg': output_msg,
    }
    return render(request, 'dashboard/question/question_create.html', context)


@login_required
@permission_required(['dashboard.view_question', 'dashboard.change_question', 'dashboard.delete_question'],
                     raise_exception=True)
def question_manage_view(request):
    # g_i_a_v = goes_in_all_view(request)
    profile = Profile.objects.get(user=request.user)
    if profile.user.is_superuser or profile.user.is_supervisor:
        return question_list_supervisor_view(request)
    all_question = Question.objects.filter(creator__user=request.user)
    if all_question.count() == 0:
        output_msg = {
            'color': 'danger',
            'text': 'سوالی برای نمایش وجود ندارد!'
        }
    else:
        output_msg = None
    context = {
        'page_name': 'سوال',
        # 'g_i_a_v': g_i_a_v,
        'output_msg': output_msg,
        'all_question': pagination_show(request, all_question),
        'profile': profile,
    }
    return render(request, 'dashboard/question/question_manage.html', context)


@login_required
@permission_required(['dashboard.view_question', 'dashboard.change_question', 'dashboard.delete_question'],
                     raise_exception=True)
def question_change_view(request, qes_id):
    # g_i_a_v = goes_in_all_view(request)
    q_instance = Question.objects.get(pk=qes_id)
    all_choice = Choice.objects.filter(choice_question=q_instance)
    if request.method == 'POST':
        question = QuestionChangeForm(request.POST, request.FILES, instance=q_instance)
        if question.is_valid():
            question.save()
            output_msg = {
                'color': 'success',
                'text': 'سوال با موفقیت ویرایش شد!'
            }
        else:
            output_msg = {
                'color': 'danger',
                'text': 'متاسفانه ویرایش سوال با مشکل مواجه شد!'
            }
    else:
        question = QuestionChangeForm(instance=q_instance)
        output_msg = None
    context = {
        'page_name': 'سوال',
        # 'g_i_a_v': g_i_a_v,
        'output_msg': output_msg,
        'all_choice': all_choice,
        'q_instance': q_instance,
        'question': question,
        'profile': Profile.objects.get(user=request.user),
    }
    return render(request, 'dashboard/question/question_change.html', context)


@login_required
@permission_required(['dashboard.view_question', 'dashboard.change_question', 'dashboard.delete_question'],
                     raise_exception=True)
def choice_change_view(request, cho_id):
    # g_i_a_v = goes_in_all_view(request)
    c_instance = Choice.objects.get(pk=cho_id)
    if request.method == 'POST':
        choice = ChoiceChangeForm(request.POST, instance=c_instance)
        if choice.is_valid():
            choice.save()
            if Choice.objects.get(pk=cho_id).is_correct:
                temp_q = Choice.objects.get(pk=cho_id).choice_question
                temp_chs = Choice.objects.filter(choice_question=temp_q)

                for i in temp_chs:
                    if i.pk != cho_id:
                        i.is_correct = False
                        i.save()

            output_msg = {
                'color': 'success',
                'text': 'گزینه با موفقیت ویرایش شد!'
            }
        else:
            output_msg = {
                'color': 'danger',
                'text': 'متاسفانه ویرایش گزینه با مشکل مواجه شد!'
            }
    else:
        choice = ChoiceChangeForm(instance=c_instance)
        output_msg = None
    context = {
        'page_name': 'گزینه',
        # 'g_i_a_v': g_i_a_v,
        'output_msg': output_msg,
        'choice': choice,
        'profile': Profile.objects.get(user=request.user),
    }
    return render(request, 'dashboard/question/choice_change.html', context)


@login_required
def question_list_view(request):
    # g_i_a_v = goes_in_all_view(request)
    profile = Profile.objects.get(user=request.user)
    if profile.user.is_supervisor:
        return question_list_supervisor_view(request)

    context = {
        # 'g_i_a_v': g_i_a_v,
        'profile': profile,
    }
    return render(request, 'dashboard/supervisor_dashboard/question_list.html', context)


@login_required
def question_list_supervisor_view(request):
    # g_i_a_v = goes_in_all_view(request)
    profile = Profile.objects.get(user=request.user)
    all_q = Question.objects.filter(creator__academy=profile.academy)
    if profile.user.is_superuser:
        all_q = Question.objects.all()
    output_msg = None
    all_out = []
    if request.method == 'POST':
        if request.POST.get('search_content'):
            all_q = all_q.filter(question_label__contains=request.POST.get('search_content'))

    for i in range(len(all_q)):
        x = Choice.objects.filter(choice_question=all_q[i])
        all_out.append({
            'question': all_q[i],
            'answers': x,
        })

    if len(all_out) == 0:
        output_msg = {
            'color': 'danger',
            'text': 'موردی برای نمایش وجود ندارد!'
        }
    all_out, page_range = pagination_show_long(request, all_out, divide_by=2)
    context = {
        'page_name': 'سوال',
        # 'g_i_a_v': g_i_a_v,
        'output_msg': output_msg,
        'all_out': all_out,
        'page_range': page_range,
        'profile': profile,
    }
    return render(request, 'dashboard/supervisor_dashboard/question_list.html', context)


# //////////////////////////////////////////// EXAMS ////////////////////////////////////////////
def examination_show(request, all_objects, page_name='page'):
    if len(all_objects) == 0:
        return {
            'color': 'danger',
            'text': ' موردی یافت نشد!',
        }
    else:
        return pagination_template_long(request, source=all_objects, page_name=page_name, divide_by=1)


@login_required
@permission_required('dashboard.add_exam', raise_exception=True)
def exam_create_view(request):
    # g_i_a_v = goes_in_all_view(request)
    profile = Profile.objects.get(user=request.user)
    if request.method == 'POST':
        exam = ExamCreateForm(request.POST, q_maker=profile)
        if exam.is_valid():
            exam.save(commit=False)
            exam.instance.exam_creator = profile
            stu_number = Profile.objects.filter(Q(user__is_student=True) & Q(klass=exam.instance.exam_klass)).count()
            temp_wallet = Wallet.objects.get(user=profile.user)
            if temp_wallet.balance >= stu_number:
                temp_wallet.spend(amount=stu_number, description='آزمون: ' + exam.instance.exam_name)
                exam.save()
                output_msg = {
                    'color': 'success',
                    'text': 'آزمون با موفقیت ایجاد شد!'
                }
            else:
                output_msg = {
                    'color': 'danger',
                    'text': 'اعتبار شما کافی نمیباشد!'
                }
        else:
            output_msg = {
                'color': 'danger',
                'text': 'متاسفانه ایجاد آزمون با مشکل مواجه شد!'
            }
    else:
        exam = ExamCreateForm(q_maker=profile)
        output_msg = None
    context = {
        'page_name': 'آزمون',
        # 'g_i_a_v': g_i_a_v,
        'exam': exam,
        'output_msg': output_msg,
        'profile': profile,
    }
    return render(request, 'dashboard/exam/exam_create.html', context)


@login_required
@permission_required(['dashboard.view_exam', 'dashboard.change_exam', 'dashboard.delete_exam'], raise_exception=True)
def exam_manage_view(request):
    # g_i_a_v = goes_in_all_view(request)
    profile = Profile.objects.get(user=request.user)
    all_exam = Exam.objects.filter(exam_creator=profile).order_by('-exam_start')
    if profile.user.is_superuser:
        all_exam = Exam.objects.all().order_by('-exam_start')
        output_msg = None
        if all_exam.count() == 0:
            output_msg = {
                'color': 'danger',
                'text': 'آزمونی برای نمایش وجود ندارد!'
            }
    elif all_exam.count() == 0:
        output_msg = {
            'color': 'danger',
            'text': 'آزمونی برای نمایش وجود ندارد!'
        }
    else:
        output_msg = None
    for i in all_exam:
        i.update_exam_status()

    all_exam, page_range = pagination_show_long(request, all_exam)
    context = {
        'page_name': 'آزمون',
        # 'g_i_a_v': g_i_a_v,
        'output_msg': output_msg,
        'all_exam': all_exam,
        'page_range': page_range,
        'profile': profile,
    }
    return render(request, 'dashboard/exam/exam_manage.html', context)


@login_required
@permission_required(['dashboard.view_exam', 'dashboard.change_exam', 'dashboard.delete_exam'], raise_exception=True)
def exam_change_view(request, exa_id):
    # g_i_a_v = goes_in_all_view(request)
    profile = Profile.objects.get(user=request.user)
    e_instance = Exam.objects.get(pk=exa_id)
    if request.method == 'POST':
        exam = ExamChangeForm(request.POST, instance=e_instance, q_maker=profile)
        if exam.is_valid():
            exam.save()
            output_msg = {
                'color': 'success',
                'text': 'آزمون با موفقیت ویرایش شد!'
            }
        else:
            output_msg = {
                'color': 'danger',
                'text': 'متاسفانه ویرایش آزمون با مشکل مواجه شد!'
            }
    else:
        exam = ExamChangeForm(instance=e_instance, q_maker=profile)
        output_msg = None
    context = {
        'page_name': 'آزمون',
        'exam': exam,
        'e_instance': e_instance,
        'output_msg': output_msg,
        # 'g_i_a_v': g_i_a_v,
        'profile': profile,
    }
    return render(request, 'dashboard/exam/exam_change.html', context)


@login_required
def exam_info_view(request, exa_id):
    # g_i_a_v = goes_in_all_view(request)
    profile = Profile.objects.get(user=request.user)
    exam = Exam.objects.get(pk=exa_id)
    resume_exam = False

    if not is_first_time(exam=exam, student=profile):
        if is_exam_time(exam) == 0:
            exam_per_student = ExamPerStudent.objects.get(student=profile, exam=exam)
            exam_per_student.update_remain_time()
            if exam_per_student.STU_remain_time.total_seconds() != 0:
                resume_exam = 'during'
            else:
                resume_exam = 'submited'
    context = {
        'page_name': 'آزمون',
        # 'g_i_a_v': g_i_a_v,
        'exam': exam,
        'resume_exam': resume_exam,
        'profile': profile,
    }
    return render(request, 'dashboard/exam/exam_info.html', context)


@login_required
def exam_start_view(request, exa_id):
    # g_i_a_v = goes_in_all_view(request)
    profile = Profile.objects.get(user=request.user)
    exam = Exam.objects.get(pk=exa_id)
    if is_first_time(exam=exam, student=profile):
        if is_start_time(exam=exam):
            exam_per_student = create_exam_per_student(exam=exam, student=profile)
            exam_per_student.STU_start = datetime.now()
            exam_per_student.calculate_finish_date_time()
            return redirect('dashboard:exam_examing', exa_id=exa_id)
        else:
            return exam_list_view(request, wrong_time=True)
    else:
        return exam_list_view(request, not_first_time=True)


@login_required
def exam_list_view(request, wrong_time=None, not_first_time=None, exam_finished=None, exam_ended=None):
    # g_i_a_v = goes_in_all_view(request)
    profile = Profile.objects.get(user=request.user)
    if profile.user.is_supervisor:
        return exam_list_supervisor_view(request)
    all_kls = profile.klass.all()
    all_exam = Exam.objects.filter(exam_klass=all_kls[0]).order_by('-exam_start')
    for i in all_exam:
        i.update_exam_status()

    if request.GET.get('exam_finished'):
        exam_finished = True
    if wrong_time:
        output_msg = {
            'color': 'warning',
            'text': 'زمان اجرای آزمون صحیح نمیباشد!'
        }
    elif exam_ended:
        output_msg = {
            'color': 'info',
            'text': 'آزمون با موفقیت خاتمه یافت!'
        }
    elif exam_finished:
        output_msg = {
            'color': 'info',
            'text': 'زمان آزمون به پایان رسید!'
        }
    elif not_first_time:
        output_msg = {
            'color': 'danger',
            'text': 'قبلا در این آزمون شرکت کرده اید!'
        }
    elif all_exam.count() == 0:
        output_msg = {
            'color': 'danger',
            'text': 'آزمونی برای نمایش وجود ندارد!'
        }
    else:
        output_msg = None

    all_exam, page_range = pagination_show_long(request, all_exam)
    context = {
        'page_name': 'آزمون',
        # 'g_i_a_v': g_i_a_v,
        'output_msg': output_msg,
        'all_exam': all_exam,
        'page_range': page_range,
        'profile': profile,
    }
    return render(request, 'dashboard/exam/exam_list.html', context)


@login_required
def exam_examing_view(request, exa_id):
    # g_i_a_v = goes_in_all_view(request)
    profile = Profile.objects.get(user=request.user)
    exam = Exam.objects.get(pk=exa_id)
    exam_per_student = ExamPerStudent.objects.get(student=profile, exam=exam)
    temp_q = exam_per_student.STU_questions
    question = []
    question_choices = []
    success_alert = False
    for i in temp_q:
        x = Question.objects.get(pk=i)
        question.append(x)
        question_choices.append(StudentAnswerForm(question=x))

    if request.method == 'POST':
        exam_per_student.save_answer(question=request.POST.get('question'), answer=request.POST.get('answer_option'))
        success_alert = True

    question_answers = exam_per_student.STU_answers
    all_objects = []

    for i in range(len(question)):
        if Choice.objects.filter(pk=question_answers[i]).count() != 0:
            temp_ch = Choice.objects.get(pk=question_answers[i])
            all_objects.append([i + 1, 'گزینه {}'.format(temp_ch.choice_label)])
        else:
            all_objects.append([i + 1, 'سفید'])

    question, page_range = examination_show(request, question, 'question')
    question_choices, page_range = examination_show(request, question_choices, 'question')
    question_answers, page_range = examination_show(request, question_answers, 'question')

    has_answered = question_answers[0]
    if Choice.objects.filter(pk=has_answered).count() != 0:
        has_answered = Choice.objects.get(pk=has_answered)
        has_answered = 'گزینه ثبت شده:' + has_answered.choice_label
    else:
        has_answered = 'پاسخی برای این سوال ثبت نکرده اید!'

    remaining_time = exam_per_student.update_remain_time()

    if remaining_time == -1:
        return exam_list_view(request, exam_finished=True)
    context = {
        # 'g_i_a_v': g_i_a_v,
        'exam': exam,
        'question': question,
        'page_range': page_range,
        'question_choices': question_choices,
        'remaining_time': remaining_time,
        'success_alert': success_alert,
        'has_answered': has_answered,
        'all_objects': all_objects,
        'profile': profile,
    }
    return render(request, 'dashboard/exam/exam_examing.html', context)


@login_required
def exam_end_view(request, exa_id):
    # g_i_a_v = goes_in_all_view(request)
    profile = Profile.objects.get(user=request.user)
    exam = Exam.objects.get(pk=exa_id)
    exam_per_student = ExamPerStudent.objects.get(student=profile, exam=exam)
    remaining_time = exam_per_student.update_remain_time()
    if request.POST.get('endit'):
        exam_per_student.STU_remain_time = 0
        exam_per_student.update_remain_time(entry=True)
        return exam_list_view(request, exam_ended=True)
    context = {
        'page_name': 'آزمون',
        # 'g_i_a_v': g_i_a_v,
        'remaining_time': remaining_time,
        'exam': exam,
        'profile': profile,
    }
    return render(request, 'dashboard/exam/exam_end.html', context)


@login_required
def exam_score_view(request):
    # g_i_a_v = goes_in_all_view(request)
    profile = Profile.objects.get(user=request.user)
    output_msg = None
    if profile.user.is_student is False:
        return supervisor_exam_klass_list(request)

    all_kls = profile.klass.all()
    all_exam = Exam.objects.filter(exam_klass=all_kls[0]).order_by('-exam_start')
    all_scores, all_eps = [], []

    for i in all_exam:
        if StudentScore.objects.filter(student=profile, exam=i).count() == 0:
            if ExamPerStudent.objects.filter(student=profile, exam=i).count() == 0:
                all_eps.append(None)
                x = None
            else:
                all_eps.append(ExamPerStudent.objects.get(student=profile, exam=i))
                x = ExamPerStudent.objects.get(student=profile, exam=i).calculate_the_score()
        else:
            x = StudentScore.objects.get(student=profile, exam=i)
            all_eps.append(ExamPerStudent.objects.get(student=profile, exam=i))
        all_scores.append(x)

    all_out = []
    for i in range(len(all_exam)):
        x = {
            'exam': all_exam[i],
            'score': all_scores[i],
            'eps': all_eps[i],
        }
        all_out.append(x)

    if len(all_out) == 0:
        output_msg = {
            'color': 'danger',
            'text': 'موردی برای نمایش وجود ندارد!'
        }
    all_out, page_range = pagination_template_long(request, all_out, divide_by=2)
    context = {
        'page_name': 'نمره',
        # 'g_i_a_v': g_i_a_v,
        'output_msg': output_msg,
        'all_out': all_out,
        'page_range': page_range,
        'profile': profile,
    }
    return render(request, 'dashboard/exam/exam_score.html', context)


@login_required
def supervisor_exam_klass_list(request):
    # g_i_a_v = goes_in_all_view(request)
    profile = Profile.objects.get(user=request.user)
    output_msg = None
    if profile.user.is_superuser:
        all_klass = Klass.objects.all()
    else:
        all_klass = Klass.objects.filter(academy=profile.academy)
    if all_klass.count() == 0:
        output_msg = {
            'color': 'danger',
            'text': 'موردی برای نمایش وجود ندارد!'
        }
    all_klass, page_range = pagination_template_long(request, all_klass, divide_by=6)
    context = {
        'page_name': 'کلاس',
        # 'g_i_a_v': g_i_a_v,
        'output_msg': output_msg,
        'all_klass': all_klass,
        'page_range': page_range,
        'profile': profile,
    }
    return render(request, 'dashboard/exam/exam_score.html', context)


@login_required
def supervisor_exam_klass_list_for_score_view(request, kls_id):
    # g_i_a_v = goes_in_all_view(request)
    output_msg = None
    klass = Klass.objects.get(pk=kls_id)
    all_exam_in_klass = Exam.objects.filter(exam_klass=klass)
    if len(all_exam_in_klass) == 0:
        output_msg = {
            'color': 'danger',
            'text': 'موردی برای نمایش وجود ندارد!'
        }
    all_exam_in_klass, page_range = pagination_template_long(request, all_exam_in_klass, divide_by=6)
    context = {
        'page_name': 'آزمون',
        # 'g_i_a_v': g_i_a_v,
        'output_msg': output_msg,
        'all_exam_in_klass': all_exam_in_klass,
        'page_range': page_range,
        'profile': Profile.objects.get(user=request.user),
    }
    return render(request, 'dashboard/exam/exam_list_score.html', context)


@login_required
def supervisor_exam_list_for_score_view(request, exa_id):
    # g_i_a_v = goes_in_all_view(request)
    profile = Profile.objects.get(user=request.user)
    output_msg = None
    all_exam = Exam.objects.filter(exam_klass__academy=profile.academy)
    for i in all_exam:
        i.update_exam_status()

    all_exam, page_range = pagination_template_long(request, all_exam, divide_by=6)
    if len(all_exam) == 0:
        output_msg = {
            'color': 'danger',
            'text': 'موردی برای نمایش وجود ندارد!'
        }
    context = {
        'page_name': 'نمره',
        # 'g_i_a_v': g_i_a_v,
        'output_msg': output_msg,
        'all_exam': all_exam,
        'page_range': page_range,
        'profile': profile,
    }
    return render(request, 'dashboard/exam/exam_list_score.html', context)


@login_required
def supervisor_exam_score_list_view(request, exa_id, message=None):
    # g_i_a_v = goes_in_all_view(request)
    output_msg = None
    if message:
        return render(request, 'dashboard/exam/exam_score_list.html',
                      {'page_name': 'نمره', 'output_msg': output_msg, 'message': message,
                       'exa_id': exa_id, })
    else:
        exam = Exam.objects.get(pk=exa_id)
        all_student = Profile.objects.filter(Q(user__is_student=True) & Q(klass=exam.exam_klass))
        if all_student.count() != 0:
            for i in all_student:
                if StudentScore.objects.filter(Q(exam=exam) & Q(student=i)).count() == 0:
                    all_eps = ExamPerStudent.objects.filter(Q(exam=exam) & Q(student=i))
                    if all_eps != 0:
                        for j in all_eps:
                            j.calculate_the_score()

        all_score = StudentScore.objects.filter(exam=exam)
        temp_now = datetime.now()

        if len(all_score) == 0:
            all_score = None
            page_range = None
            output_msg = {
                'color': 'danger',
                'text': 'موردی برای نمایش وجود ندارد!'
            }
        elif exam.exam_finish >= temp_now:
            all_score = None
            page_range = None
            output_msg = {
                'color': 'danger',
                'text': 'هنوز زمان آزمون به پایان نرسیده است!'
            }
        else:
            all_score, page_range = pagination_show_long(request, all_score, divide_by=6)

    context = {
        'page_name': 'نمره',
        # 'g_i_a_v': g_i_a_v,
        'output_msg': output_msg,
        'all_score': all_score,
        'page_range': page_range,
        'exa_id': exa_id,
        'message': message,
        'profile': Profile.objects.get(user=request.user),
    }
    return render(request, 'dashboard/exam/exam_score_list.html', context)


@login_required
def supervisor_exam_score_for_student_list_view(request, exa_id, student):
    # g_i_a_v = goes_in_all_view(request)
    output_msg = None
    exam = Exam.objects.get(pk=exa_id)
    the_student = ExamPerStudent.objects.get(student=student, exam__exam_id=exa_id)
    page_name = 'پاسخ های دانش آموز: {}'.format(the_student.student.user.get_full_name())

    all_question = list([Question.objects.get(pk=i), 'dark'] for i in the_student.STU_questions)
    right_answer = []
    for i in the_student.STU_questions:
        temp_x = Choice.objects.filter(choice_question_id=i)
        for j in temp_x:
            if j.is_correct:
                right_answer.append([j.choice_label, 'dark'])

    all_answer = []
    for i in the_student.STU_answers:
        if Choice.objects.filter(choice_id=i).count() == 1:
            tmep_z = []
            tmep_z.append(Choice.objects.get(choice_id=i).choice_label)
            if Choice.objects.get(choice_id=i).is_correct:
                tmep_z.append('success')
            else:
                tmep_z.append('danger')
            all_answer.append(tmep_z)
        else:
            all_answer.append(['بدون پاسخ', 'warning'])

    all_score = list([all_question[i], right_answer[i], all_answer[i]] for i in range(len(all_question)))

    context = {
        'page_name': page_name,
        # 'g_i_a_v': g_i_a_v,
        'output_msg': output_msg,
        'all_score': all_score,
        'exa_id': exa_id,
        'profile': Profile.objects.get(user=request.user),
    }
    return render(request, 'dashboard/exam/exam_score_per_student.html', context)


@login_required
def supervisor_exam_score_excel_view(request, exa_id):
    exam = Exam.objects.get(pk=exa_id)
    temp_now = datetime.now()

    if exam.exam_finish >= temp_now:
        return HttpResponse('هنوز زمان آزمون به اتمام نرسیده است!')

    output = io.BytesIO()
    all_student = Profile.objects.filter(Q(user__is_student=True) & Q(klass=exam.exam_klass))
    if all_student.count() != 0:
        for i in all_student:
            if StudentScore.objects.filter(Q(exam=exam) & Q(student=i)).count() == 0:
                all_eps = ExamPerStudent.objects.filter(Q(exam=exam) & Q(student=i))
                if all_eps != 0:
                    for j in all_eps:
                        j.calculate_the_score()

    all_score = StudentScore.objects.filter(exam=exam)

    workbook = xlsxwriter.Workbook(output)
    worksheet = workbook.add_worksheet()

    center = workbook.add_format({'align': 'center'})

    worksheet.set_column('A:A', 30)
    worksheet.set_column('B:B', 12)

    temp_list = ['نام و نام خانوادگی', 'کد ملی', 'کل پاسخ ها', 'نمره', 'درصد']

    for i in range(5):
        worksheet.write(0, i, temp_list[i], center)

    for i in range(len(all_score)):
        worksheet.write(i + 1, 0, all_score[i].student.user.get_full_name(), center)
        worksheet.write(i + 1, 1, all_score[i].student.passport_number, center)
        worksheet.write(i + 1, 2, all_score[i].all_questions, center)
        worksheet.write(i + 1, 3, all_score[i].score, center)
        worksheet.write(i + 1, 4, all_score[i].percentage, center)

    # Close the workbook before sending the data.
    workbook.close()

    output.seek(0)
    filename = 'output.xlsx'
    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=%s' % filename

    return response


@login_required
def exam_list_supervisor_view(request):
    # g_i_a_v = goes_in_all_view(request)
    profile = Profile.objects.get(user=request.user)
    output_msg = None
    all_out = Exam.objects.filter(exam_klass__academy=profile.academy).order_by('-exam_finish')
    for i in all_out:
        i.update_exam_status()

    if len(all_out) == 0:
        output_msg = {
            'color': 'danger',
            'text': 'موردی برای نمایش وجود ندارد!'
        }
    all_out, page_range = pagination_show_long(request, all_out, divide_by=2)
    context = {
        'page_name': 'آزمون',
        # 'g_i_a_v': g_i_a_v,
        'output_msg': output_msg,
        'all_out': all_out,
        'page_range': page_range,
        'profile': profile,
    }
    return render(request, 'dashboard/supervisor_dashboard/exam_list.html', context)


# //////////////////////////////////////////// DELETE ////////////////////////////////////////////
@login_required
def delete_objects_view(request, name, obj_id):
    # g_i_a_v = goes_in_all_view(request)
    if name.lower() == 'notification':
        temp_obj = Notification.objects.get(pk=obj_id)
        if temp_obj and request.GET.get('status'):
            temp_obj.delete()
            output_msg = {
                'color': 'success',
                'text': 'اطلاعیه با موفقیت حذف شد!'
            }
        else:
            output_msg = None
    elif name.lower() == 'category':
        temp_obj = Category.objects.get(pk=obj_id)
        if temp_obj and request.GET.get('status'):
            temp_cats = Category.objects.filter(parent=temp_obj)
            for i in temp_cats:
                i.delete()
            temp_obj.delete()
            output_msg = {
                'color': 'success',
                'text': 'دسته بندی با موفقیت حذف شد!'
            }
        else:
            output_msg = None
    elif name.lower() == 'klass':
        temp_obj = Klass.objects.get(pk=obj_id)
        if temp_obj and request.GET.get('status'):
            temp_stus = User.objects.filter(Q(profile__klass=temp_obj) & Q(is_student=True))
            if temp_stus.count() > 0:
                for i in temp_stus:
                    i.delete()
            temp_teachers = Profile.objects.filter(Q(klass=temp_obj) & Q(user__is_teacher=True))
            if temp_teachers.count() > 0:
                for i in temp_teachers:
                    i.klass.remove(temp_obj)
                    i.save()
            if SignUpKey.objects.filter(klass=temp_obj).count() != 0:
                temp_key = SignUpKey.objects.get(klass=temp_obj)
                temp_key.delete()
            if Exam.objects.filter(exam_klass=temp_obj).count() != 0:
                temp_exams = Exam.objects.filter(exam_klass=temp_obj)
                for i in temp_exams:
                    i.delete()
            temp_obj.delete()
            output_msg = {
                'color': 'success',
                'text': 'کلاس با موفقیت حذف شد!'
            }
        else:
            output_msg = None
    elif name.lower() == 'academy':
        temp_obj = Academy.objects.get(pk=obj_id)
        if temp_obj and request.GET.get('status'):
            temp_users = User.objects.filter(profile__academy=temp_obj)
            if temp_users.count() > 0:
                for i in temp_users:
                    i.delete()
            temp_klss = Klass.objects.filter(academy=temp_obj)
            if temp_klss.count() > 0:
                for i in temp_klss:
                    temp_key = SignUpKey.objects.get(klass=i)
                    temp_key.delete()
                    i.delete()
            temp_obj.delete()
            output_msg = {
                'color': 'success',
                'text': 'آموزشگاه با موفقیت حذف شد!'
            }
        else:
            output_msg = None
    elif name.lower() == 'user':
        temp_obj = User.objects.get(pk=obj_id)
        if temp_obj and request.GET.get('status'):
            temp_obj.delete()
            output_msg = {
                'color': 'success',
                'text': 'کاربر با موفقیت حذف شد!'
            }
        else:
            output_msg = None
    else:
        output_msg = None

    # output_msg = None
    context = {
        'output_msg': output_msg,
        # 'g_i_a_v': g_i_a_v,
        'temp_obj': temp_obj,
        'profile': Profile.objects.get(user=request.user),
    }
    return render(request, 'dashboard/delete/delete_message.html', context)


@login_required
def delete_uuid_objects_view(request, name, obj_uuid):
    # g_i_a_v = goes_in_all_view(request)
    if name.lower() == 'question':
        temp_obj = Question.objects.get(pk=obj_uuid)
        temp_cho = Choice.objects.filter(choice_question_id=obj_uuid)
        if temp_obj and request.GET.get('status'):
            for i in temp_cho:
                i.delete()
            temp_obj.delete()
            output_msg = {
                'color': 'success',
                'text': 'سوال با موفقیت حذف شد!'
            }
        else:
            output_msg = None
    elif name.lower() == 'exam':
        temp_obj = Exam.objects.get(pk=obj_uuid)
        if temp_obj and request.GET.get('status'):
            temp_obj.delete()
            output_msg = {
                'color': 'success',
                'text': 'آزمون با موفقیت حذف شد!'
            }
        else:
            output_msg = None
    else:
        output_msg = None
    context = {
        'output_msg': output_msg,
        # 'g_i_a_v': g_i_a_v,
        'temp_obj': temp_obj,
        'profile': Profile.objects.get(user=request.user),
    }
    return render(request, 'dashboard/delete/delete_message.html', context)


# //////////////////////////////////////////// USER MANAGEMENT ////////////////////////////////////////////
@login_required
@permission_required(
    ['dashboard.view_user', 'dashboard.change_user', 'dashboard.delete_user', 'dashboard.delete_category'],
    raise_exception=True)
def list_supervisor_view(request):
    # g_i_a_v = goes_in_all_view(request)
    profile = Profile.objects.get(user=request.user)
    if profile.user.is_superuser:
        all_supervisor = Profile.objects.filter(user__is_supervisor=True)
    else:
        all_supervisor = Question.objects.filter(creator__academy=profile.academy)
    output_msg = None
    if request.method == 'POST':
        if request.POST.get('search_content'):
            all_supervisor = all_supervisor.filter(Q(passport_number__exact=request.POST.get('search_content')) | Q(
                user__first_name__contains=request.POST.get('search_content')) | Q(
                user__last_name__contains=request.POST.get('search_content')))
    if all_supervisor.count() == 0:
        output_msg = {
            'color': 'danger',
            'text': 'سرپرستی با این مشخصات یافت نشد!',
        }
    all_supervisor, page_range = pagination_template_long(request, all_supervisor, divide_by=2)
    context = {
        'page_name': 'سرپرست',
        # 'g_i_a_v': g_i_a_v,
        'output_msg': output_msg,
        'all_supervisor': all_supervisor,
        'page_range': page_range,
        'profile': profile,
    }
    return render(request, 'dashboard/superuser_dashboard/supervisor_list.html', context)


@login_required
@permission_required(['dashboard.view_user', 'dashboard.change_user', 'dashboard.delete_user'],
                     raise_exception=True)
def list_teacher_view(request):
    # g_i_a_v = goes_in_all_view(request)
    profile = Profile.objects.get(user=request.user)
    if profile.user.is_superuser:
        all_teacher = Profile.objects.filter(user__is_teacher=True)
    else:
        all_teacher = Profile.objects.filter(Q(user__is_teacher=True) & Q(academy=profile.academy))
    output_msg = None
    if request.method == 'POST':
        if request.POST.get('search_content'):
            all_teacher = all_teacher.filter(Q(passport_number__exact=request.POST.get('search_content')) | Q(
                user__first_name__contains=request.POST.get('search_content')) | Q(
                user__last_name__contains=request.POST.get('search_content')))
    if all_teacher.count() == 0:
        output_msg = {
            'color': 'danger',
            'text': 'آموزگاری با این مشخصات یافت نشد!',
        }
    all_teacher, page_range = pagination_template_long(request, all_teacher, divide_by=2)
    context = {
        # 'g_i_a_v': g_i_a_v,
        'output_msg': output_msg,
        'page_name': 'آموزگار',
        'all_teacher': all_teacher,
        'page_range': page_range,
        'profile': profile,
    }
    return render(request, 'dashboard/superuser_dashboard/teacher_list.html', context)


@login_required
@permission_required(['dashboard.view_user', 'dashboard.change_user', 'dashboard.delete_user'],
                     raise_exception=True)
def list_student_view(request):
    # g_i_a_v = goes_in_all_view(request)
    profile = Profile.objects.get(user=request.user)
    if profile.user.is_superuser:
        all_student = Profile.objects.filter(user__is_student=True)
    else:
        all_student = Profile.objects.filter(Q(academy=profile.academy) & Q(user__is_student=True))
    output_msg = None
    if request.method == 'POST':
        if request.POST.get('search_content'):
            all_student = all_student.filter(Q(passport_number__exact=request.POST.get('search_content')) | Q(
                user__first_name__contains=request.POST.get('search_content')) | Q(
                user__last_name__contains=request.POST.get('search_content')))
    if all_student.count() == 0:
        output_msg = {
            'color': 'danger',
            'text': 'دانش آموزی با این مشخصات یافت نشد!',
        }
    all_student, page_range = pagination_template_long(request, all_student, divide_by=2)
    context = {
        'page_name': 'دانش آموز',
        # 'g_i_a_v': g_i_a_v,
        'output_msg': output_msg,
        'all_student': all_student,
        'page_range': page_range,
        'profile': profile,
    }
    return render(request, 'dashboard/superuser_dashboard/student_list.html', context)


@login_required
@permission_required(['dashboard.view_user', 'dashboard.change_user', 'dashboard.delete_user'],
                     raise_exception=True)
def list_all_view(request, name):
    # g_i_a_v = goes_in_all_view(request)
    all_obj = None
    output_msg = None
    profile = Profile.objects.get(user=request.user)
    if name == 'klass':
        page_name = 'کلاس'
        if profile.user.is_superuser:
            all_obj = Klass.objects.all()
        else:
            all_obj = Klass.objects.filter(academy=profile.academy)
    elif name == 'academy':
        page_name = 'آموزشگاه'
        if profile.user.is_superuser:
            all_obj = Academy.objects.all()
    elif name == 'student':
        page_name = 'دانش آموز'
        if profile.user.is_superuser:
            all_obj = Profile.objects.filter(user__is_student=True)
        else:
            all_obj = Profile.objects.filter(Q(academy=profile.academy) & Q(user__is_student=True))
    elif name == 'teacher':
        page_name = 'آموزگار'
        if profile.user.is_superuser:
            all_obj = Profile.objects.filter(user__is_teacher=True)
        else:
            all_obj = Profile.objects.filter(Q(academy=profile.academy) & Q(user__is_teacher=True))
    elif name == 'supervisor':
        page_name = 'سرپرست'
        if profile.user.is_superuser:
            all_obj = Profile.objects.filter(user__is_supervisor=True)
        else:
            all_obj = Question.objects.filter(creator__academy=profile.academy)
    elif name == 'question':
        page_name = 'سوال'
        if profile.user.is_superuser:
            all_obj = Question.objects.all()
        else:
            all_obj = Question.objects.filter(creator__academy=profile.academy)
    else:
        raise Http404()

    if all_obj.count() == 0:
        output_msg = {
            'color': 'danger',
            'text': 'موردی یافت نشد!',
        }

    all_obj, page_range = pagination_template_long(request, all_obj, divide_by=30)
    context = {
        'name': name,
        'page_name': page_name,
        # 'g_i_a_v': g_i_a_v,
        'output_msg': output_msg,
        'all_obj': all_obj,
        'page_range': page_range,
        'profile': profile,
    }
    return render(request, 'dashboard/superuser_dashboard/all_list.html', context)


# //////////////////////////////////////////// WALLETS ///////////////////////////////////////////
@login_required
@permission_required(['dashboard.view_wallet', 'dashboard.view_transaction', 'dashboard.add_transaction'],
                     raise_exception=True)
def wallet_increase_view(request, wal_id):
    # g_i_a_v = goes_in_all_view(request)
    profile = Profile.objects.get(user=request.user)
    if profile.user.is_supervisor:
        return wallet_increase_supervisor_view(request, wal_id)
    temp_wallet = Wallet.objects.get(pk=wal_id)
    if request.method == 'POST':
        temp_amount = WalletIncreaseForm(request.POST)
        if temp_amount.is_valid():
            temp_wallet.deposit(amount=temp_amount.cleaned_data['amount'])
            temp_wallet.save()
            output_msg = {
                'color': 'success',
                'text': 'اعتبار با موفقیت افزوده شد!'
            }
        else:
            output_msg = None
    else:
        temp_amount = WalletIncreaseForm()
        output_msg = None
    context = {
        # 'g_i_a_v': g_i_a_v,
        'temp_wallet': temp_wallet,
        'output_msg': output_msg,
        'temp_amount': temp_amount,
        'profile': profile,
    }
    return render(request, 'dashboard/wallet/increase.html', context)


@login_required
@permission_required(
    ['dashboard.view_transaction', 'dashboard.add_transaction', 'dashboard.view_wallet'],
    raise_exception=True)
def wallet_transaction_list_view(request):
    # g_i_a_v = goes_in_all_view(request)
    all_transaction = Transaction.objects.filter(wallet__user=request.user).order_by('-order_time')
    output_msg = None
    if all_transaction.count() == 0:
        output_msg = {
            'color': 'danger',
            'text': 'موردی یافت نشد!',
        }
    context = {
        'page_name': 'تراکنش',
        # 'g_i_a_v': g_i_a_v,
        'output_msg': output_msg,
        'all_transaction': pagination_show(request, all_transaction, divide_by=7),
        'profile': Profile.objects.get(user=request.user),
    }
    return render(request, 'dashboard/wallet/transaction_list.html', context)


@login_required
@permission_required(
    ['dashboard.view_transaction', 'dashboard.add_transaction', 'dashboard.view_wallet'],
    raise_exception=True)
def wallet_increase_supervisor_view(request, wal_id):
    # g_i_a_v = goes_in_all_view(request)
    profile = Profile.objects.get(user=request.user)
    self_wallet = Wallet.objects.get(user=profile.user)
    teacher_wallet = Wallet.objects.get(pk=wal_id)
    if request.method == 'POST':
        temp_amount = WalletIncreaseForm(request.POST)
        if temp_amount.is_valid():
            if self_wallet.transfer_to_teacher(amount=temp_amount.cleaned_data['amount'], other_user=teacher_wallet):
                output_msg = {
                    'color': 'success',
                    'text': 'اعتبار با موفقیت افزوده شد!'
                }
            else:
                output_msg = {
                    'color': 'danger',
                    'text': 'موجودی شما کافی نیست!'
                }
        else:
            output_msg = None
    else:
        temp_amount = WalletIncreaseForm()
        output_msg = None
    context = {
        # 'g_i_a_v': g_i_a_v,
        'temp_wallet': teacher_wallet,
        'self_wallet': self_wallet,
        'output_msg': output_msg,
        'temp_amount': temp_amount,
        'profile': profile,
    }
    return render(request, 'dashboard/wallet/increase.html', context)


@login_required
@permission_required(
    ['dashboard.delete_user'],
    raise_exception=True)
def changepass_user_view(request, user_id, name):
    # g_i_a_v = goes_in_all_view(request)
    output_msg, change_pass, temp_prof = None, None, None
    if name == 'user':
        temp_user = User.objects.get(pk=user_id)
        temp_prof = Profile.objects.get(user=temp_user)
        if request.method == 'POST':
            change_pass = ChangePasswordForm(request.POST)
            if change_pass.is_valid():
                temp_pass = change_pass.cleaned_data['password']
                temp_pass = make_password(temp_pass)
                temp_user.password = temp_pass
                temp_user.save()
                output_msg = {
                    'color': 'success',
                    'text': 'تغییر رمز با موفقیت انجام شد!',
                }
            else:
                output_msg = {
                    'color': 'danger',
                    'text': 'تغییر رمز با مشکل مواجه شد!',
                }
        change_pass = ChangePasswordForm()
    else:
        output_msg = {
            'color': 'danger',
            'text': 'درخواست صحیح نمیباشد!',
        }
    context = {
        # 'g_i_a_v': g_i_a_v,
        'output_msg': output_msg,
        'change_pass': change_pass,
        'temp_prof': temp_prof,
        'profile': Profile.objects.get(user=request.user),
    }
    return render(request, 'dashboard/delete/changepass_message.html', context)


# //////////////////////////////////////////// COPIES ///////////////////////////////////////////
@login_required
@permission_required(
    ['dashboard.delete_notification'], raise_exception=True)
def copy_object_view(request, name):
    # g_i_a_v = goes_in_all_view(request)
    errors, output_msg, all_obj, page_name = None, None, None, None
    if name == 'exam':
        page_name = 'آزمون'
        if request.method == 'POST':
            all_obj = CopyObjectForm(request.POST)
            if all_obj.is_valid():
                exam = all_obj.cleaned_data['exam']
                klass = all_obj.cleaned_data['klass']
                exam.update_exam_status()
                if exam.is_active:
                    new_exam = exam
                    new_exam.pk = None
                    new_exam.save()
                    new_exam.exam_klass = klass
                    new_exam.save()
                    output_msg = {
                        'color': 'success',
                        'text': 'آزمون با موفقیت کپی شد!',
                    }
                else:
                    output_msg = {
                        'color': 'danger',
                        'text': 'زمان آزمون به پایان رسیده است!',
                    }

        else:
            all_obj = CopyObjectForm()
    else:
        pass
    context = {
        # 'g_i_a_v': g_i_a_v,
        'page_name': page_name,
        'all_obj': all_obj,
        'name': name,
        'output_msg': output_msg,
        'errors': errors,
        'profile': Profile.objects.get(user=request.user),
    }
    return render(request, 'dashboard/superuser_dashboard/copy_object.html', context)


# //////////////////////////////////////////// IMPORTS ///////////////////////////////////////////
@login_required
@permission_required(
    ['dashboard.view_user', 'dashboard.delete_user'], raise_exception=True)
def import_choose_view(request):
    # g_i_a_v = goes_in_all_view(request)
    profile = Profile.objects.get(user=request.user)
    output_msg, klass, academy = None, None, None
    if profile.user.is_supervisor:
        return redirect('dashboard:import_student', aca_id=profile.academy.pk)
    elif request.method == 'POST':
        academy = AcademyFinderForm(request.POST)
        if academy.is_valid():
            academy = academy.cleaned_data['academy']
            return redirect('dashboard:import_student', aca_id=academy.pk)
    else:
        academy = AcademyFinderForm()
    context = {
        # 'g_i_a_v': g_i_a_v,
        'output_msg': output_msg,
        'academy': academy,
        'klass': klass,
        'profile': profile,
    }
    return render(request, 'dashboard/superuser_dashboard/import_choose.html', context)


@login_required
@permission_required(
    ['dashboard.view_user', 'dashboard.delete_user'], raise_exception=True)
def import_student_view(request, aca_id):
    # g_i_a_v = goes_in_all_view(request)
    academy = Academy.objects.get(pk=aca_id)
    output_msg, errors = None, []
    if request.method == 'POST':
        klass = ImportStudentForm(request.POST, request.FILES, academy=academy)
        if klass.is_valid():
            excel_file = klass.cleaned_data['exl_file']
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.sheetnames
            sheet = wb[sheet[0]]
            for i in range(1, len(sheet['A'])):
                pass_num = sheet['A' + str(i)].value
                f_name = sheet['B' + str(i)].value
                l_name = sheet['C' + str(i)].value
                if Profile.objects.filter(passport_number__exact=pass_num).count() == 0:
                    pswd = make_password(str(pass_num))
                    try:
                        temp_user = User.objects.create(is_student=True, first_name=f_name, last_name=l_name,
                                                        username=pass_num,
                                                        password=pswd)
                        temp_prof = Profile.objects.get(user=temp_user.pk)
                        temp_prof.passport_number = pass_num
                        temp_prof.academy = academy
                        temp_prof.klass.add(klass.cleaned_data['klass'])
                        temp_prof.save()
                    except:
                        pass
                else:
                    errors.append({
                        'pass_num': pass_num,
                        'f_name': f_name,
                        'l_name': l_name,
                    })

            output_msg = {
                'color': 'success',
                'text': 'افزودن گروهی دانش آموزان با موفقیت انجام شد!',
            }
        else:
            output_msg = {
                'color': 'danger',
                'text': 'متاسفانه افزودن گروهی دانش آموزان با مشکل مواجه شد!',
            }
    else:
        klass = ImportStudentForm(academy=academy)
    context = {
        # 'g_i_a_v': g_i_a_v,
        'output_msg': output_msg,
        'errors': errors,
        'klass': klass,
        'profile': Profile.objects.get(user=request.user),
    }
    return render(request, 'dashboard/superuser_dashboard/import_student.html', context)


# //////////////////////////////////////////// CHECKERS ///////////////////////////////////////////

@login_required
@permission_required(
    ['dashboard.view_user', 'dashboard.delete_user'], raise_exception=True)
def extra_exam_checker(request):
    all_scores = StudentScore.objects.all()
    all_exams = []
    all_student = []
    for i in all_scores:
        if i.exam not in all_exams:
            all_exams.append(i.exam)
        if i.student not in all_student:
            all_student.append(i.student)
    for i in all_exams:
        for j in all_student:
            if StudentScore.objects.filter(Q(exam=i) & Q(student=j)).count() > 1:
                temp_x = StudentScore.objects.filter(Q(exam=i) & Q(student=j))
                for k in range(1, len(temp_x)):
                    temp_x[k].delete()

    return HttpResponse('Done')


@login_required
@permission_required(
    ['dashboard.view_user', 'dashboard.delete_user'], raise_exception=True)
def extra_score_checker(request, exa_id):
    # g_i_a_v = goes_in_all_view(request)
    if request.user.is_superuser:
        all_eps = ExamPerStudent.objects.filter(exam_id=exa_id)
        for i in all_eps:
            i.re_calculate_the_score()

        return supervisor_exam_score_list_view(request, exa_id, message='Done')
    else:
        return HttpResponseRedirect(reverse('home_page'))


@login_required
@permission_required(
    ['dashboard.view_user', 'dashboard.delete_user'], raise_exception=True)
def extra_score_checker_student(request, stu_id):
    # g_i_a_v = goes_in_all_view(request)
    profile = Profile.objects.get(user=request.user)
    if profile.user.is_superuser:
        page_tit = 'محاسبه مجدد تمامی نمرات '
        page_tit += str(Profile.objects.get(user_id=stu_id))
        all_eps = ExamPerStudent.objects.filter(student__user_id=stu_id)
        for i in all_eps:
            i.re_calculate_the_score()
        output = 'student-recalculate'
        output_msg = {
            'color': 'success',
            'text': 'محاسبه مجدد تمامی نمرات با موفقیت انجام شد'
        }
        context = {
            # 'g_i_a_v': g_i_a_v,
            'page_tit': page_tit,
            'output': output,
            'output_msg': output_msg,
            'profile': profile,
        }
        return render(request, 'dashboard/superuser_dashboard/messages.html', context)
    else:
        return HttpResponseRedirect(reverse('home_page'))
